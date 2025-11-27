#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
F&F 원가 대시보드 - 아이템별 원가율 Excel 파일 생성 스크립트

사용 방법:
    python generate_item_cost_rate_csv.py --period 26SS
    python generate_item_cost_rate_csv.py --period 25FW
    python generate_item_cost_rate_csv.py --period 25SS

이 스크립트는 지정된 기간의 모든 브랜드 CSV 파일을 읽어서
아이템별 원가율 데이터를 Excel 파일로 생성합니다.
각 브랜드별로 시트가 구분되어 있으며, 당년-전년 비교 데이터를 포함합니다.

핵심 원칙:
- 원부자재 = 원자재 + 부자재 + 본사공급자재 + 택/라벨 (아트웍 제외!)
- 원가율 = (평균원가 ÷ (평균TAG / 1.1)) × 100
- 수량 가중 평균 사용
- USD와 KRW 별도 계산
- 환율: 브랜드-시즌-중분류 조합으로 FX.csv에서 조회
- 전시즌 TAG USD 변환: 전시즌 환율 사용
"""

import pandas as pd
import argparse
import os
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# FX 파일 경로
FX_FILE = 'public/COST RAW/FX.csv'

# 브랜드 코드 매핑
BRAND_MAP = {
    'M': 'MLB',
    'I': 'MLB KIDS',
    'X': 'DISCOVERY',
    'ST': 'SERGIO TACCHINI',
    'V': 'DUVETICA'
}

# 카테고리 → FX 카테고리 매핑
def map_category_to_fx_category(category: str) -> str:
    """중분류 카테고리를 의류/슈즈/용품으로 매핑"""
    if pd.isna(category) or category == '':
        return '의류'
    
    category_upper = str(category).strip().upper()
    
    # 의류: Outer, Inner, Bottom, Wear_etc
    if category_upper in ['OUTER', 'INNER', 'BOTTOM', 'WEAR_ETC']:
        return '의류'
    
    # 슈즈: Shoes
    if category_upper == 'SHOES':
        return '슈즈'
    
    # 용품: Bag, Headwear, Acc_etc
    if category_upper in ['BAG', 'HEADWEAR', 'ACC_ETC', 'ACC']:
        return '용품'
    
    # 기본값: 의류
    return '의류'

# FX.csv에서 환율 조회
def get_exchange_rate(df_fx: pd.DataFrame, brand_code: str, season_code: str, category: str = None) -> float:
    """FX.csv에서 특정 브랜드/시즌/카테고리의 환율 조회"""
    if df_fx is None or df_fx.empty:
        return 1300.0
    
    fx_category = map_category_to_fx_category(category) if category else '의류'
    
    filtered = df_fx[
        (df_fx['브랜드'] == brand_code) &
        (df_fx['시즌'] == season_code) &
        (df_fx['카테고리'] == fx_category)
    ]
    
    if len(filtered) > 0:
        rate = float(filtered.iloc[0]['환율'])
        if rate > 0:
            return rate
    
    # 기본값: 의류 환율 사용
    if category and fx_category != '의류':
        default = df_fx[
            (df_fx['브랜드'] == brand_code) &
            (df_fx['시즌'] == season_code) &
            (df_fx['카테고리'] == '의류')
        ]
        if len(default) > 0:
            rate = float(default.iloc[0]['환율'])
            if rate > 0:
                return rate
    
    return 1300.0

# 전시즌 코드 계산
def get_previous_season(season: str) -> str:
    """전시즌 코드 반환"""
    season_map = {
        '26SS': '25SS',
        '26S': '25S',
        '25SS': '24SS',
        '25S': '24S',
        '25FW': '24FW',
        '25F': '24F',
        '24SS': '23SS',
        '24S': '23S',
        '24FW': '23FW',
        '24F': '23F',
    }
    return season_map.get(season, '')

# 시즌 코드 변환 (26SS → 26S, 25SS → 25S)
def convert_season_format(season: str) -> str:
    """시즌 형식 변환"""
    if season.endswith('SS'):
        return season.replace('SS', 'S')
    if season.endswith('FW'):
        return season.replace('FW', 'F')
    return season

# 시즌 정규화 (25SS, 25S → 25S)
def normalize_season(season_str):
    """시즌 문자열 정규화"""
    if pd.isna(season_str):
        return ''
    s = str(season_str).strip().upper()
    if s.endswith('SS'):
        return s[:-1]  # 25SS → 25S
    if s.endswith('FW'):
        return s[:-1]  # 25FW → 25F
    return s

# 기간 폴더명 결정
def get_season_folder(season: str) -> str:
    """기간 폴더명 결정"""
    season_upper = season.upper()
    if season_upper in ['26SS', '26S']:
        return '26SS'
    elif season_upper in ['25SS', '25S']:
        return '25S'
    elif season_upper in ['25FW', '25F']:
        return '25FW'
    elif season_upper in ['24SS', '24S']:
        return '24S'
    elif season_upper in ['24FW', '24F']:
        return '24FW'
    else:
        return season_upper

# 아이템별 집계 계산
def aggregate_by_item(df: pd.DataFrame, df_fx: pd.DataFrame, brand_code: str, 
                     current_season_code: str, prev_season_code: str) -> pd.DataFrame:
    """아이템별로 데이터 집계"""
    
    # 전년/당년 시즌 필터링
    df_prev = df[df.iloc[:, 1].apply(normalize_season).isin([prev_season_code, prev_season_code + 'S', prev_season_code + 'SS'])]
    df_curr = df[df.iloc[:, 1].apply(normalize_season).isin([current_season_code, current_season_code + 'S', current_season_code + 'SS'])]
    
    # 아이템별 그룹핑 (중분류 + 아이템명)
    item_map = {}
    
    # 전년 데이터 집계
    for idx, row in df_prev.iterrows():
        category = str(row.iloc[3]).strip() if len(row) > 3 else ''
        item_name = str(row.iloc[4]).strip() if len(row) > 4 else ''
        key = f"{category}_{item_name}"
        
        if key not in item_map:
            item_map[key] = {
                'category': category,
                'item_name': item_name,
                'brand': brand_code,
                'prev_data': [],
                'curr_data': []
            }
        
        item_map[key]['prev_data'].append(row)
    
    # 당년 데이터 집계
    for idx, row in df_curr.iterrows():
        category = str(row.iloc[3]).strip() if len(row) > 3 else ''
        item_name = str(row.iloc[4]).strip() if len(row) > 4 else ''
        key = f"{category}_{item_name}"
        
        if key not in item_map:
            item_map[key] = {
                'category': category,
                'item_name': item_name,
                'brand': brand_code,
                'prev_data': [],
                'curr_data': []
            }
        
        item_map[key]['curr_data'].append(row)
    
    # 아이템별 계산
    results = []
    
    for key, item_data in item_map.items():
        prev_rows = item_data['prev_data']
        curr_rows = item_data['curr_data']
        
        # 전년 데이터 계산
        qty_prev = sum(pd.to_numeric(row.iloc[7], errors='coerce') or 0 for row in prev_rows)
        
        # 전년 TAG (KRW)
        tag_prev_krw_total = sum((pd.to_numeric(row.iloc[6], errors='coerce') or 0) * 
                                 (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in prev_rows)
        avg_tag_prev_krw = tag_prev_krw_total / qty_prev if qty_prev > 0 else 0
        
        # 전년 TAG (USD) - 전시즌 환율 사용
        tag_prev_usd_total = 0
        for row in prev_rows:
            category = row.iloc[3] if len(row) > 3 else None
            fx_rate = get_exchange_rate(df_fx, brand_code, prev_season_code, category)
            tag_krw = pd.to_numeric(row.iloc[6], errors='coerce') or 0
            qty_row = pd.to_numeric(row.iloc[7], errors='coerce') or 0
            tag_prev_usd_total += (tag_krw / fx_rate) * qty_row
        avg_tag_prev_usd = tag_prev_usd_total / qty_prev if qty_prev > 0 else 0
        
        # 전년 원가 (USD) - 총금액 컬럼 사용
        material_prev_total = sum(pd.to_numeric(row.iloc[30], errors='coerce') or 0 for row in prev_rows)
        artwork_prev_total = sum(pd.to_numeric(row.iloc[31], errors='coerce') or 0 for row in prev_rows)
        labor_prev_total = sum(pd.to_numeric(row.iloc[32], errors='coerce') or 0 for row in prev_rows)
        margin_prev_total = sum(pd.to_numeric(row.iloc[33], errors='coerce') or 0 for row in prev_rows)
        expense_prev_total = sum(pd.to_numeric(row.iloc[34], errors='coerce') or 0 for row in prev_rows)
        
        material_prev = material_prev_total / qty_prev if qty_prev > 0 else 0
        artwork_prev = artwork_prev_total / qty_prev if qty_prev > 0 else 0
        labor_prev = labor_prev_total / qty_prev if qty_prev > 0 else 0
        margin_prev = margin_prev_total / qty_prev if qty_prev > 0 else 0
        expense_prev = expense_prev_total / qty_prev if qty_prev > 0 else 0
        avg_cost_prev_usd = material_prev + artwork_prev + labor_prev + margin_prev + expense_prev
        
        # 전년 원가율 (USD) - 세부 항목별
        tag_excl_vat_prev_usd = avg_tag_prev_usd / 1.1 if avg_tag_prev_usd > 0 else 0
        cost_rate_material_prev_usd = (material_prev / tag_excl_vat_prev_usd) * 100 if tag_excl_vat_prev_usd > 0 else 0
        cost_rate_artwork_prev_usd = (artwork_prev / tag_excl_vat_prev_usd) * 100 if tag_excl_vat_prev_usd > 0 else 0
        cost_rate_labor_prev_usd = (labor_prev / tag_excl_vat_prev_usd) * 100 if tag_excl_vat_prev_usd > 0 else 0
        cost_rate_margin_prev_usd = (margin_prev / tag_excl_vat_prev_usd) * 100 if tag_excl_vat_prev_usd > 0 else 0
        cost_rate_expense_prev_usd = (expense_prev / tag_excl_vat_prev_usd) * 100 if tag_excl_vat_prev_usd > 0 else 0
        cost_rate_prev_usd = cost_rate_material_prev_usd + cost_rate_artwork_prev_usd + cost_rate_labor_prev_usd + cost_rate_margin_prev_usd + cost_rate_expense_prev_usd
        
        # 전년 원가 (KRW)
        material_prev_krw_total = sum((pd.to_numeric(row.iloc[22], errors='coerce') or 0) * 
                                     (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in prev_rows)
        sub_prev_krw_total = sum((pd.to_numeric(row.iloc[24], errors='coerce') or 0) * 
                                 (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in prev_rows)
        hq_prev_krw_total = sum((pd.to_numeric(row.iloc[27], errors='coerce') or 0) * 
                               (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in prev_rows)
        tag_label_prev_krw_total = sum((pd.to_numeric(row.iloc[25], errors='coerce') or 0) * 
                                       (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in prev_rows)
        artwork_prev_krw_total = sum((pd.to_numeric(row.iloc[23], errors='coerce') or 0) * 
                                     (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in prev_rows)
        labor_prev_krw_total = sum((pd.to_numeric(row.iloc[26], errors='coerce') or 0) * 
                                   (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in prev_rows)
        margin_prev_krw_total = sum((pd.to_numeric(row.iloc[28], errors='coerce') or 0) * 
                                   (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in prev_rows)
        expense_prev_krw_total = sum((pd.to_numeric(row.iloc[29], errors='coerce') or 0) * 
                                     (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in prev_rows)
        
        material_prev_krw = (material_prev_krw_total + sub_prev_krw_total + hq_prev_krw_total + tag_label_prev_krw_total) / qty_prev if qty_prev > 0 else 0
        artwork_prev_krw = artwork_prev_krw_total / qty_prev if qty_prev > 0 else 0
        labor_prev_krw = labor_prev_krw_total / qty_prev if qty_prev > 0 else 0
        margin_prev_krw = margin_prev_krw_total / qty_prev if qty_prev > 0 else 0
        expense_prev_krw = expense_prev_krw_total / qty_prev if qty_prev > 0 else 0
        avg_cost_prev_krw = material_prev_krw + artwork_prev_krw + labor_prev_krw + margin_prev_krw + expense_prev_krw
        
        # 전년 원가율 (KRW) - 세부 항목별
        tag_excl_vat_prev_krw = avg_tag_prev_krw / 1.1 if avg_tag_prev_krw > 0 else 0
        cost_rate_material_prev_krw = (material_prev_krw / tag_excl_vat_prev_krw) * 100 if tag_excl_vat_prev_krw > 0 else 0
        cost_rate_artwork_prev_krw = (artwork_prev_krw / tag_excl_vat_prev_krw) * 100 if tag_excl_vat_prev_krw > 0 else 0
        cost_rate_labor_prev_krw = (labor_prev_krw / tag_excl_vat_prev_krw) * 100 if tag_excl_vat_prev_krw > 0 else 0
        cost_rate_margin_prev_krw = (margin_prev_krw / tag_excl_vat_prev_krw) * 100 if tag_excl_vat_prev_krw > 0 else 0
        cost_rate_expense_prev_krw = (expense_prev_krw / tag_excl_vat_prev_krw) * 100 if tag_excl_vat_prev_krw > 0 else 0
        cost_rate_prev_krw = cost_rate_material_prev_krw + cost_rate_artwork_prev_krw + cost_rate_labor_prev_krw + cost_rate_margin_prev_krw + cost_rate_expense_prev_krw
        
        # 당년 데이터 계산
        qty_curr = sum(pd.to_numeric(row.iloc[7], errors='coerce') or 0 for row in curr_rows)
        
        # 당년 TAG (KRW)
        tag_curr_krw_total = sum((pd.to_numeric(row.iloc[6], errors='coerce') or 0) * 
                                 (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in curr_rows)
        avg_tag_curr_krw = tag_curr_krw_total / qty_curr if qty_curr > 0 else 0
        
        # 당년 TAG (USD) - 전시즌 환율 사용
        tag_curr_usd_total = 0
        for row in curr_rows:
            category = row.iloc[3] if len(row) > 3 else None
            fx_rate = get_exchange_rate(df_fx, brand_code, prev_season_code, category)
            tag_krw = pd.to_numeric(row.iloc[6], errors='coerce') or 0
            qty_row = pd.to_numeric(row.iloc[7], errors='coerce') or 0
            tag_curr_usd_total += (tag_krw / fx_rate) * qty_row
        avg_tag_curr_usd = tag_curr_usd_total / qty_curr if qty_curr > 0 else 0
        
        # 당년 원가 (USD) - 총금액 컬럼 사용
        material_curr_total = sum(pd.to_numeric(row.iloc[30], errors='coerce') or 0 for row in curr_rows)
        artwork_curr_total = sum(pd.to_numeric(row.iloc[31], errors='coerce') or 0 for row in curr_rows)
        labor_curr_total = sum(pd.to_numeric(row.iloc[32], errors='coerce') or 0 for row in curr_rows)
        margin_curr_total = sum(pd.to_numeric(row.iloc[33], errors='coerce') or 0 for row in curr_rows)
        expense_curr_total = sum(pd.to_numeric(row.iloc[34], errors='coerce') or 0 for row in curr_rows)
        
        material_curr = material_curr_total / qty_curr if qty_curr > 0 else 0
        artwork_curr = artwork_curr_total / qty_curr if qty_curr > 0 else 0
        labor_curr = labor_curr_total / qty_curr if qty_curr > 0 else 0
        margin_curr = margin_curr_total / qty_curr if qty_curr > 0 else 0
        expense_curr = expense_curr_total / qty_curr if qty_curr > 0 else 0
        avg_cost_curr_usd = material_curr + artwork_curr + labor_curr + margin_curr + expense_curr
        
        # 당년 원가율 (USD) - 세부 항목별
        tag_excl_vat_curr_usd = avg_tag_curr_usd / 1.1 if avg_tag_curr_usd > 0 else 0
        cost_rate_material_curr_usd = (material_curr / tag_excl_vat_curr_usd) * 100 if tag_excl_vat_curr_usd > 0 else 0
        cost_rate_artwork_curr_usd = (artwork_curr / tag_excl_vat_curr_usd) * 100 if tag_excl_vat_curr_usd > 0 else 0
        cost_rate_labor_curr_usd = (labor_curr / tag_excl_vat_curr_usd) * 100 if tag_excl_vat_curr_usd > 0 else 0
        cost_rate_margin_curr_usd = (margin_curr / tag_excl_vat_curr_usd) * 100 if tag_excl_vat_curr_usd > 0 else 0
        cost_rate_expense_curr_usd = (expense_curr / tag_excl_vat_curr_usd) * 100 if tag_excl_vat_curr_usd > 0 else 0
        cost_rate_curr_usd = cost_rate_material_curr_usd + cost_rate_artwork_curr_usd + cost_rate_labor_curr_usd + cost_rate_margin_curr_usd + cost_rate_expense_curr_usd
        
        # 당년 원가 (KRW)
        material_curr_krw_total = sum((pd.to_numeric(row.iloc[22], errors='coerce') or 0) * 
                                     (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in curr_rows)
        sub_curr_krw_total = sum((pd.to_numeric(row.iloc[24], errors='coerce') or 0) * 
                                 (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in curr_rows)
        hq_curr_krw_total = sum((pd.to_numeric(row.iloc[27], errors='coerce') or 0) * 
                               (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in curr_rows)
        tag_label_curr_krw_total = sum((pd.to_numeric(row.iloc[25], errors='coerce') or 0) * 
                                       (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in curr_rows)
        artwork_curr_krw_total = sum((pd.to_numeric(row.iloc[23], errors='coerce') or 0) * 
                                     (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in curr_rows)
        labor_curr_krw_total = sum((pd.to_numeric(row.iloc[26], errors='coerce') or 0) * 
                                   (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in curr_rows)
        margin_curr_krw_total = sum((pd.to_numeric(row.iloc[28], errors='coerce') or 0) * 
                                   (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in curr_rows)
        expense_curr_krw_total = sum((pd.to_numeric(row.iloc[29], errors='coerce') or 0) * 
                                     (pd.to_numeric(row.iloc[7], errors='coerce') or 0) for row in curr_rows)
        
        material_curr_krw = (material_curr_krw_total + sub_curr_krw_total + hq_curr_krw_total + tag_label_curr_krw_total) / qty_curr if qty_curr > 0 else 0
        artwork_curr_krw = artwork_curr_krw_total / qty_curr if qty_curr > 0 else 0
        labor_curr_krw = labor_curr_krw_total / qty_curr if qty_curr > 0 else 0
        margin_curr_krw = margin_curr_krw_total / qty_curr if qty_curr > 0 else 0
        expense_curr_krw = expense_curr_krw_total / qty_curr if qty_curr > 0 else 0
        avg_cost_curr_krw = material_curr_krw + artwork_curr_krw + labor_curr_krw + margin_curr_krw + expense_curr_krw
        
        # 당년 원가율 (KRW) - 세부 항목별
        tag_excl_vat_curr_krw = avg_tag_curr_krw / 1.1 if avg_tag_curr_krw > 0 else 0
        cost_rate_material_curr_krw = (material_curr_krw / tag_excl_vat_curr_krw) * 100 if tag_excl_vat_curr_krw > 0 else 0
        cost_rate_artwork_curr_krw = (artwork_curr_krw / tag_excl_vat_curr_krw) * 100 if tag_excl_vat_curr_krw > 0 else 0
        cost_rate_labor_curr_krw = (labor_curr_krw / tag_excl_vat_curr_krw) * 100 if tag_excl_vat_curr_krw > 0 else 0
        cost_rate_margin_curr_krw = (margin_curr_krw / tag_excl_vat_curr_krw) * 100 if tag_excl_vat_curr_krw > 0 else 0
        cost_rate_expense_curr_krw = (expense_curr_krw / tag_excl_vat_curr_krw) * 100 if tag_excl_vat_curr_krw > 0 else 0
        cost_rate_curr_krw = cost_rate_material_curr_krw + cost_rate_artwork_curr_krw + cost_rate_labor_curr_krw + cost_rate_margin_curr_krw + cost_rate_expense_curr_krw
        
        results.append({
            '중분류': item_data['category'],
            '아이템명': item_data['item_name'],
            '브랜드': brand_code,
            # USD 원가율 (세부 항목별)
            'USD원가율_재료계_전년': round(cost_rate_material_prev_usd, 1),
            'USD원가율_재료계_당년': round(cost_rate_material_curr_usd, 1),
            'USD원가율_아트웍_전년': round(cost_rate_artwork_prev_usd, 1),
            'USD원가율_아트웍_당년': round(cost_rate_artwork_curr_usd, 1),
            'USD원가율_공임_전년': round(cost_rate_labor_prev_usd, 1),
            'USD원가율_공임_당년': round(cost_rate_labor_curr_usd, 1),
            'USD원가율_마진_전년': round(cost_rate_margin_prev_usd, 1),
            'USD원가율_마진_당년': round(cost_rate_margin_curr_usd, 1),
            'USD원가율_경비_전년': round(cost_rate_expense_prev_usd, 1),
            'USD원가율_경비_당년': round(cost_rate_expense_curr_usd, 1),
            'USD원가율_합계_전년': round(cost_rate_prev_usd, 1),
            'USD원가율_합계_당년': round(cost_rate_curr_usd, 1),
            # USD 평균단가 (세부 항목별)
            'USD평균단가_재료계_전년': round(material_prev, 2),
            'USD평균단가_재료계_당년': round(material_curr, 2),
            'USD평균단가_아트웍_전년': round(artwork_prev, 2),
            'USD평균단가_아트웍_당년': round(artwork_curr, 2),
            'USD평균단가_공임_전년': round(labor_prev, 2),
            'USD평균단가_공임_당년': round(labor_curr, 2),
            'USD평균단가_마진_전년': round(margin_prev, 2),
            'USD평균단가_마진_당년': round(margin_curr, 2),
            'USD평균단가_경비_전년': round(expense_prev, 2),
            'USD평균단가_경비_당년': round(expense_curr, 2),
            'USD평균단가_합계_전년': round(avg_cost_prev_usd, 2),
            'USD평균단가_합계_당년': round(avg_cost_curr_usd, 2),
            # KRW 원가율 (세부 항목별)
            'KRW원가율_재료계_전년': round(cost_rate_material_prev_krw, 1),
            'KRW원가율_재료계_당년': round(cost_rate_material_curr_krw, 1),
            'KRW원가율_아트웍_전년': round(cost_rate_artwork_prev_krw, 1),
            'KRW원가율_아트웍_당년': round(cost_rate_artwork_curr_krw, 1),
            'KRW원가율_공임_전년': round(cost_rate_labor_prev_krw, 1),
            'KRW원가율_공임_당년': round(cost_rate_labor_curr_krw, 1),
            'KRW원가율_마진_전년': round(cost_rate_margin_prev_krw, 1),
            'KRW원가율_마진_당년': round(cost_rate_margin_curr_krw, 1),
            'KRW원가율_경비_전년': round(cost_rate_expense_prev_krw, 1),
            'KRW원가율_경비_당년': round(cost_rate_expense_curr_krw, 1),
            'KRW원가율_합계_전년': round(cost_rate_prev_krw, 1),
            'KRW원가율_합계_당년': round(cost_rate_curr_krw, 1),
            # 기타
            '수량_전년': int(qty_prev),
            '수량_당년': int(qty_curr),
            '평균TAG_KRW_전년': round(avg_tag_prev_krw, 0),
            '평균TAG_KRW_당년': round(avg_tag_curr_krw, 0),
            'TAG금액_전년': round(tag_prev_krw_total, 0),
            'TAG금액_당년': round(tag_curr_krw_total, 0),
        })
    
    return pd.DataFrame(results)

# 발주비중 계산
def calculate_order_share(df_items: pd.DataFrame) -> pd.DataFrame:
    """발주비중 계산 (TAG 금액 기준, 브랜드별로 계산)"""
    # 브랜드별로 그룹핑하여 각 브랜드 내에서 발주비중 계산
    df_items['발주비중_전년'] = 0.0
    df_items['발주비중_당년'] = 0.0
    
    for brand_code in df_items['브랜드'].unique():
        brand_mask = df_items['브랜드'] == brand_code
        df_brand = df_items[brand_mask]
        
        # 브랜드별 TAG 금액 합계
        brand_total_tag_prev = df_brand['TAG금액_전년'].sum()
        brand_total_tag_curr = df_brand['TAG금액_당년'].sum()
        
        # 브랜드 내 아이템별 발주비중 계산
        df_items.loc[brand_mask, '발주비중_전년'] = (
            df_brand['TAG금액_전년'] / brand_total_tag_prev * 100
        ) if brand_total_tag_prev > 0 else 0
        
        df_items.loc[brand_mask, '발주비중_당년'] = (
            df_brand['TAG금액_당년'] / brand_total_tag_curr * 100
        ) if brand_total_tag_curr > 0 else 0
    
    return df_items

# Excel 파일 생성
def create_excel_file(df_all: pd.DataFrame, output_file: str, period: str):
    """Excel 파일 생성 (USD 원가율, USD 평균단가, KRW 원가율 시트 분리)"""
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # 발주비중 계산
    df_all = calculate_order_share(df_all)
    
    # 중분류 순서 정의
    category_order = ['Outer', 'Inner', 'Bottom', 'Acc_etc', 'Wear_etc']
    
    # 1. USD 원가율 시트
    create_cost_rate_sheet(wb, df_all, 'USD 원가율', 'USD원가율', category_order)
    
    # 2. USD 평균단가 시트
    create_avg_cost_sheet(wb, df_all, 'USD 평균단가', 'USD평균단가', category_order)
    
    # 3. KRW 원가율 시트
    create_cost_rate_sheet(wb, df_all, 'KRW 원가율', 'KRW원가율', category_order, include_tag=True)
    
    # 파일 저장
    wb.save(output_file)
    print(f"[OK] Excel 파일 생성 완료: {output_file}")

def create_cost_rate_sheet(wb: Workbook, df_all: pd.DataFrame, sheet_name: str, 
                           prefix: str, category_order: list, include_tag: bool = False):
    """원가율 시트 생성"""
    ws = wb.create_sheet(title=sheet_name)
    
    # 컬럼 정의
    columns = [
        '중분류', '아이템명', '브랜드',
        f'{prefix}_재료계_전년', f'{prefix}_재료계_당년',
        f'{prefix}_아트웍_전년', f'{prefix}_아트웍_당년',
        f'{prefix}_공임_전년', f'{prefix}_공임_당년',
        f'{prefix}_마진_전년', f'{prefix}_마진_당년',
        f'{prefix}_경비_전년', f'{prefix}_경비_당년',
        f'{prefix}_합계_전년', f'{prefix}_합계_당년',
        '발주비중_전년', '발주비중_당년',
        '수량_전년', '수량_당년'
    ]
    
    if include_tag:
        columns.extend(['평균TAG_KRW_전년', '평균TAG_KRW_당년'])
    
    # 헤더 작성
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    row_idx = 2
    brands = ['M', 'I', 'X', 'ST', 'V']
    
    for brand_code in brands:
        df_brand = df_all[df_all['브랜드'] == brand_code].copy()
        if len(df_brand) == 0:
            continue
        
        # 중분류별로 정렬
        df_brand['중분류_순서'] = df_brand['중분류'].apply(
            lambda x: category_order.index(x) if x in category_order else 999
        )
        df_brand = df_brand.sort_values(['중분류_순서', '아이템명'])
        df_brand = df_brand.drop('중분류_순서', axis=1)
        
        # 중분류별로 그룹핑하여 처리
        categories = df_brand['중분류'].unique()
        
        for category in categories:
            df_category = df_brand[df_brand['중분류'] == category].copy()
            
            # 데이터 작성
            for _, row_data in df_category.iterrows():
                # 데이터 행 작성
                for col_idx, col_name in enumerate(columns, 1):
                    value = row_data.get(col_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                
                row_idx += 1
            
            # 중분류 소계 추가
            add_subtotal_row(ws, row_idx, columns, df_category, prefix, include_tag, df_all)
            row_idx += 1
        
        # 브랜드별 합계
        add_brand_total_row(ws, row_idx, columns, df_brand, prefix, include_tag, 
                           BRAND_MAP.get(brand_code, brand_code), df_all)
        row_idx += 2  # 빈 행 추가
    
    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

def create_avg_cost_sheet(wb: Workbook, df_all: pd.DataFrame, sheet_name: str, 
                          prefix: str, category_order: list):
    """평균단가 시트 생성"""
    ws = wb.create_sheet(title=sheet_name)
    
    # 컬럼 정의
    columns = [
        '중분류', '아이템명', '브랜드',
        f'{prefix}_재료계_전년', f'{prefix}_재료계_당년',
        f'{prefix}_아트웍_전년', f'{prefix}_아트웍_당년',
        f'{prefix}_공임_전년', f'{prefix}_공임_당년',
        f'{prefix}_마진_전년', f'{prefix}_마진_당년',
        f'{prefix}_경비_전년', f'{prefix}_경비_당년',
        f'{prefix}_합계_전년', f'{prefix}_합계_당년',
        '발주비중_전년', '발주비중_당년',
        '수량_전년', '수량_당년'
    ]
    
    # 헤더 작성
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    row_idx = 2
    brands = ['M', 'I', 'X', 'ST', 'V']
    
    for brand_code in brands:
        df_brand = df_all[df_all['브랜드'] == brand_code].copy()
        if len(df_brand) == 0:
            continue
        
        # 중분류별로 정렬
        df_brand['중분류_순서'] = df_brand['중분류'].apply(
            lambda x: category_order.index(x) if x in category_order else 999
        )
        df_brand = df_brand.sort_values(['중분류_순서', '아이템명'])
        df_brand = df_brand.drop('중분류_순서', axis=1)
        
        # 중분류별로 그룹핑하여 처리
        categories = df_brand['중분류'].unique()
        
        for category in categories:
            df_category = df_brand[df_brand['중분류'] == category].copy()
            
            # 데이터 작성
            for _, row_data in df_category.iterrows():
                # 데이터 행 작성
                for col_idx, col_name in enumerate(columns, 1):
                    value = row_data.get(col_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                
                row_idx += 1
            
            # 중분류 소계 추가 (평균단가 시트는 include_tag=False)
            add_subtotal_row(ws, row_idx, columns, df_category, prefix, False, df_all)
            row_idx += 1
        
        # 브랜드별 합계
        add_brand_total_row(ws, row_idx, columns, df_brand, prefix, False, 
                           BRAND_MAP.get(brand_code, brand_code), df_all)
        row_idx += 2  # 빈 행 추가
    
    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

def add_subtotal_row(ws, row_idx: int, columns: list, df_subset: pd.DataFrame, 
                     prefix: str, include_tag: bool, df_all: pd.DataFrame = None):
    """중분류별 소계 행 추가"""
    # 소계 행 작성
    for col_idx, col_name in enumerate(columns, 1):
        if col_name == '중분류':
            value = f"{df_subset.iloc[0]['중분류']} 소계"
        elif col_name == '아이템명':
            value = ''
        elif col_name == '브랜드':
            value = df_subset.iloc[0]['브랜드']
        elif col_name.startswith(prefix):
            # 원가율/평균단가는 가중 평균 계산 (수량 기준)
            if '전년' in col_name:
                total_qty = df_subset['수량_전년'].sum()
                if total_qty > 0:
                    # 가중 평균: (값 × 수량) 합계 / 수량 합계
                    weighted_sum = (df_subset[col_name] * df_subset['수량_전년']).sum()
                    value = round(weighted_sum / total_qty, 2) if '평균단가' in prefix else round(weighted_sum / total_qty, 1)
                else:
                    value = ''
            else:  # 당년
                total_qty = df_subset['수량_당년'].sum()
                if total_qty > 0:
                    weighted_sum = (df_subset[col_name] * df_subset['수량_당년']).sum()
                    value = round(weighted_sum / total_qty, 2) if '평균단가' in prefix else round(weighted_sum / total_qty, 1)
                else:
                    value = ''
        elif col_name in ['발주비중_전년', '발주비중_당년']:
            # 발주비중은 브랜드별 TAG 금액 기준으로 재계산
            brand_code = df_subset.iloc[0]['브랜드']
            if df_all is not None:
                # 해당 브랜드의 전체 TAG 금액
                df_brand = df_all[df_all['브랜드'] == brand_code]
                brand_total_tag = df_brand['TAG금액_전년'].sum() if '전년' in col_name else df_brand['TAG금액_당년'].sum()
                # 소계의 TAG 금액
                subset_tag = df_subset['TAG금액_전년'].sum() if '전년' in col_name else df_subset['TAG금액_당년'].sum()
                value = round((subset_tag / brand_total_tag * 100), 2) if brand_total_tag > 0 else 0
            else:
                value = ''
        elif col_name in ['수량_전년', '수량_당년']:
            value = int(df_subset[col_name].sum()) if col_name in df_subset.columns else ''
        elif include_tag and col_name in ['평균TAG_KRW_전년', '평균TAG_KRW_당년']:
            # 가중 평균 계산
            qty_col = '수량_전년' if '전년' in col_name else '수량_당년'
            total_qty = df_subset[qty_col].sum()
            if total_qty > 0:
                tag_total = (df_subset[col_name] * df_subset[qty_col]).sum()
                value = round(tag_total / total_qty, 0) if total_qty > 0 else ''
            else:
                value = ''
        else:
            value = ''
        
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

def add_brand_total_row(ws, row_idx: int, columns: list, df_brand: pd.DataFrame, 
                        prefix: str, include_tag: bool, brand_name: str, df_all: pd.DataFrame = None):
    """브랜드별 합계 행 추가"""
    # 합계 행 작성
    for col_idx, col_name in enumerate(columns, 1):
        if col_name == '중분류':
            value = f"{brand_name} 합계"
        elif col_name == '아이템명':
            value = ''
        elif col_name == '브랜드':
            value = df_brand.iloc[0]['브랜드']
        elif col_name.startswith(prefix):
            # 원가율/평균단가는 가중 평균 계산 (수량 기준)
            if '전년' in col_name:
                total_qty = df_brand['수량_전년'].sum()
                if total_qty > 0:
                    weighted_sum = (df_brand[col_name] * df_brand['수량_전년']).sum()
                    value = round(weighted_sum / total_qty, 2) if '평균단가' in prefix else round(weighted_sum / total_qty, 1)
                else:
                    value = ''
            else:  # 당년
                total_qty = df_brand['수량_당년'].sum()
                if total_qty > 0:
                    weighted_sum = (df_brand[col_name] * df_brand['수량_당년']).sum()
                    value = round(weighted_sum / total_qty, 2) if '평균단가' in prefix else round(weighted_sum / total_qty, 1)
                else:
                    value = ''
        elif col_name in ['발주비중_전년', '발주비중_당년']:
            # 발주비중은 브랜드별 TAG 금액 기준으로 재계산 (브랜드 합계는 항상 100%)
            brand_tag = df_brand['TAG금액_전년'].sum() if '전년' in col_name else df_brand['TAG금액_당년'].sum()
            value = 100.0 if brand_tag > 0 else 0
        elif col_name in ['수량_전년', '수량_당년']:
            value = int(df_brand[col_name].sum()) if col_name in df_brand.columns else ''
        elif include_tag and col_name in ['평균TAG_KRW_전년', '평균TAG_KRW_당년']:
            # 가중 평균 계산
            qty_col = '수량_전년' if '전년' in col_name else '수량_당년'
            total_qty = df_brand[qty_col].sum()
            if total_qty > 0:
                tag_total = (df_brand[col_name] * df_brand[qty_col]).sum()
                value = round(tag_total / total_qty, 0) if total_qty > 0 else ''
            else:
                value = ''
        else:
            value = ''
        
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

def main():
    parser = argparse.ArgumentParser(description='아이템별 원가율 Excel 파일 생성')
    parser.add_argument('--period', type=str, required=True,
                       help='기간 코드 (예: 26SS, 25SS, 25FW)')
    args = parser.parse_args()
    
    period = args.period.upper()
    
    print("F&F Cost Dashboard - 아이템별 원가율 Excel 파일 생성")
    print("=" * 60)
    print(f"\n기간: {period}")
    
    # 기간 폴더명 결정
    season_folder = get_season_folder(period)
    
    # 시즌 코드 결정
    if period in ['26SS', '26S']:
        season_code = '26S'
        prev_season_code = '25S'
    elif period in ['25SS', '25S']:
        season_code = '25S'
        prev_season_code = '24S'
    elif period in ['25FW', '25F']:
        season_code = '25F'
        prev_season_code = '24F'
    elif period in ['24SS', '24S']:
        season_code = '24S'
        prev_season_code = '23S'
    elif period in ['24FW', '24F']:
        season_code = '24F'
        prev_season_code = '23F'
    else:
        season_code = convert_season_format(period)
        prev_season = get_previous_season(period)
        prev_season_code = convert_season_format(prev_season) if prev_season else ''
    
    print(f"시즌 코드: {season_code}")
    print(f"전년 시즌 코드: {prev_season_code}")
    print(f"기간 폴더: {season_folder}")
    
    # FX 파일 로드
    if not os.path.exists(FX_FILE):
        print(f"\n[ERROR] {FX_FILE} 파일이 없습니다.")
        return
    
    df_fx = pd.read_csv(FX_FILE, encoding='utf-8-sig')
    print(f"\n[OK] FX 파일 로드 완료: {len(df_fx)}개 환율 데이터")
    
    # 모든 브랜드 데이터 통합
    all_items = []
    brands = ['M', 'I', 'X', 'ST', 'V']
    
    for brand_code in brands:
        print(f"\n{'=' * 60}")
        print(f"브랜드 {brand_code} ({BRAND_MAP.get(brand_code, brand_code)}) 처리 중...")
        print(f"{'=' * 60}")
        
        # 파일명 결정
        file_season = season_code if season_code in ['25F', '26F', '24F'] else period
        csv_file = f'public/COST RAW/{season_folder}/{brand_code}_{file_season}.csv'
        
        print(f"CSV 파일: {csv_file}")
        
        # CSV 파일 존재 확인
        if not os.path.exists(csv_file):
            print(f"[WARN] {csv_file} 파일이 없습니다. 건너뜁니다.")
            continue
        
        # CSV 파일 로드
        print(f"[1] CSV 파일 로드 중...")
        df = pd.read_csv(csv_file, encoding='utf-8-sig')
        
        # 중분류 통합
        def normalize_category(category):
            if pd.isna(category):
                return 'Acc_etc'
            category_str = str(category).strip()
            category_upper = category_str.upper()
            if category_upper in ['SHOES', 'BAG', 'HEADWEAR', 'ACC_ETC', 'ACC']:
                return 'Acc_etc'
            return category_str
        
        df.iloc[:, 3] = df.iloc[:, 3].apply(normalize_category)
        
        # 수량 컬럼 클렌징
        df.iloc[:, 7] = pd.to_numeric(df.iloc[:, 7].astype(str).str.replace(',', '').str.strip(), errors='coerce').fillna(0)
        
        # USD/KRW 단가 컬럼도 숫자로 변환
        for col_idx in range(14, 35):
            if col_idx < len(df.columns):
                df.iloc[:, col_idx] = pd.to_numeric(df.iloc[:, col_idx], errors='coerce').fillna(0)
        
        print(f"   > 총 {len(df)}개 레코드 로드")
        
        # 아이템별 집계
        print(f"[2] 아이템별 집계 계산 중...")
        df_items = aggregate_by_item(df, df_fx, brand_code, season_code, prev_season_code)
        print(f"   > {len(df_items)}개 아이템 집계 완료")
        
        all_items.append(df_items)
    
    # 모든 브랜드 데이터 통합
    if len(all_items) == 0:
        print("\n[ERROR] 처리할 데이터가 없습니다.")
        return
    
    df_all = pd.concat(all_items, ignore_index=True)
    print(f"\n[OK] 전체 {len(df_all)}개 아이템 데이터 통합 완료")
    
    # Excel 파일 생성
    output_file = f'public/COST RAW/{season_folder}/item_cost_rate_{period}.xlsx'
    print(f"\n[3] Excel 파일 생성 중: {output_file}")
    create_excel_file(df_all, output_file, period)
    
    print("\n" + "=" * 60)
    print("모든 작업이 완료되었습니다!")
    print(f"출력 파일: {output_file}")
    print("=" * 60)

if __name__ == '__main__':
    main()

